from flask import Flask, render_template, request, jsonify, send_file
from flask_cors import CORS
from flask_sqlalchemy import SQLAlchemy
import pandas as pd
from datetime import datetime
import io
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

app = Flask(__name__)
CORS(app)

# Database configuration
DATABASE_URL = os.environ.get('DATABASE_URL')
if DATABASE_URL:
    if DATABASE_URL.startswith('postgres://'):
        DATABASE_URL = DATABASE_URL.replace('postgres://', 'postgresql://')
else:
    # Use SQLite as fallback when DATABASE_URL is not set
    DATABASE_URL = 'sqlite:///pnl_tracker.db'

app.config['SQLALCHEMY_DATABASE_URI'] = DATABASE_URL
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)

# Database Models
class Entry(db.Model):
    __tablename__ = 'entries'
    
    id = db.Column(db.Integer, primary_key=True)
    date = db.Column(db.Date, nullable=False)
    gain = db.Column(db.Float, default=0)
    loss = db.Column(db.Float, default=0)
    withdrawal = db.Column(db.Float, default=0)
    deposit = db.Column(db.Float, default=0)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    def to_dict(self):
        return {
            'id': self.id,
            'Date': self.date.strftime('%Y-%m-%d'),
            'Gain': self.gain,
            'Loss': self.loss,
            'Withdrawal': self.withdrawal,
            'Deposit': self.deposit
        }

class Settings(db.Model):
    __tablename__ = 'settings'
    
    id = db.Column(db.Integer, primary_key=True)
    starting_balance = db.Column(db.Float, default=0)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    
    def to_dict(self):
        return {
            'starting_balance': self.starting_balance
        }

# Create tables
with app.app_context():
    db.create_all()
    # Initialize settings if not exists
    if Settings.query.count() == 0:
        default_settings = Settings(starting_balance=0)
        db.session.add(default_settings)
        db.session.commit()

def get_starting_balance():
    """Get current starting balance from settings"""
    settings = Settings.query.first()
    return settings.starting_balance if settings else 0

def calculate_daily_metrics(data, starting_balance):
    """Calculate all derived columns for daily data"""
    if not data:
        return []
    
    df = pd.DataFrame(data)
    df['Date'] = pd.to_datetime(df['Date'])
    df = df.sort_values('Date').reset_index(drop=True)
    
    # Fill NaN with 0 for calculations
    df['Gain'] = df['Gain'].fillna(0)
    df['Loss'] = df['Loss'].fillna(0)
    df['Withdrawal'] = df['Withdrawal'].fillna(0)
    df['Deposit'] = df['Deposit'].fillna(0)
    
    # Calculate cumulative gains and losses
    df['Cgain'] = df['Gain'].cumsum()
    df['Closs'] = df['Loss'].cumsum()
    
    # Calculate daily net
    df['Net'] = df['Gain'] - df['Loss']
    
    # Calculate cumulative net
    df['Cum'] = df['Net'].cumsum()
    
    # Calculate cumulative withdrawals and deposits
    df['CWithdrawal'] = df['Withdrawal'].cumsum()
    df['CDeposit'] = df['Deposit'].cumsum()
    
    # Calculate balance using Formula B
    # Balance = Starting Balance + Cumulative Net + Cumulative Deposits - Cumulative Withdrawals
    df['Balance'] = starting_balance + df['Cum'] + df['CDeposit'] - df['CWithdrawal']
    
    # Add serial number
    df.insert(0, 'Sl', range(1, len(df) + 1))
    
    # Convert to dict for JSON
    df['Date'] = df['Date'].dt.strftime('%Y-%m-%d')
    return df.to_dict('records')

def calculate_monthly_summary(daily_data, starting_balance):
    """Calculate monthly summary from daily data"""
    if not daily_data:
        return []
    
    df = pd.DataFrame(daily_data)
    df['Date'] = pd.to_datetime(df['Date'])
    df['Month'] = df['Date'].dt.to_period('M')
    
    # Group by month
    monthly = df.groupby('Month').agg({
        'Gain': 'sum',
        'Loss': 'sum',
        'Withdrawal': 'sum',
        'Deposit': 'sum'
    }).reset_index()
    
    # Convert period to timestamp
    monthly['Month'] = monthly['Month'].dt.to_timestamp()
    
    # Calculate net
    monthly['Net'] = monthly['Gain'] - monthly['Loss']
    
    # Calculate cumulative net
    monthly['Cum'] = monthly['Net'].cumsum()
    
    # Calculate cumulative withdrawals and deposits
    monthly['CWithdrawal'] = monthly['Withdrawal'].cumsum()
    monthly['CDeposit'] = monthly['Deposit'].cumsum()
    
    # Calculate balance using Formula B
    monthly['Balance'] = starting_balance + monthly['Cum'] + monthly['CDeposit'] - monthly['CWithdrawal']
    
    # Add serial number
    monthly.insert(0, 'Sl', range(1, len(monthly) + 1))
    
    # Convert to dict for JSON
    monthly['Month'] = monthly['Month'].dt.strftime('%Y-%m-%d')
    return monthly.to_dict('records')

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/settings', methods=['GET'])
def get_settings():
    """Get current settings"""
    try:
        settings = Settings.query.first()
        return jsonify({
            'success': True,
            'data': settings.to_dict() if settings else {'starting_balance': 0}
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 400

@app.route('/api/settings', methods=['PUT'])
def update_settings():
    """Update settings"""
    try:
        data = request.json
        settings = Settings.query.first()
        
        if settings:
            settings.starting_balance = float(data.get('starting_balance', 0))
        else:
            settings = Settings(starting_balance=float(data.get('starting_balance', 0)))
            db.session.add(settings)
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Settings updated successfully'
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 400

@app.route('/api/entries', methods=['GET'])
def get_entries():
    """Get all entries from database"""
    try:
        entries = Entry.query.order_by(Entry.date).all()
        data = [entry.to_dict() for entry in entries]
        return jsonify({
            'success': True,
            'data': data
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 400

@app.route('/api/entries', methods=['POST'])
def add_entry():
    """Add new entry to database"""
    try:
        data = request.json
        entry = Entry(
            date=datetime.strptime(data['Date'], '%Y-%m-%d').date(),
            gain=float(data.get('Gain', 0)),
            loss=float(data.get('Loss', 0)),
            withdrawal=float(data.get('Withdrawal', 0)),
            deposit=float(data.get('Deposit', 0))
        )
        db.session.add(entry)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Entry added successfully'
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 400

@app.route('/api/entries/<int:entry_id>', methods=['DELETE'])
def delete_entry(entry_id):
    """Delete entry from database"""
    try:
        entry = Entry.query.get(entry_id)
        if entry:
            db.session.delete(entry)
            db.session.commit()
            return jsonify({
                'success': True,
                'message': 'Entry deleted successfully'
            })
        else:
            return jsonify({
                'success': False,
                'error': 'Entry not found'
            }), 404
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 400

@app.route('/api/calculate', methods=['POST'])
def calculate():
    """Calculate metrics for the provided data"""
    try:
        data = request.json.get('data', [])
        starting_balance = get_starting_balance()
        daily_metrics = calculate_daily_metrics(data, starting_balance)
        monthly_summary = calculate_monthly_summary(daily_metrics, starting_balance)
        
        return jsonify({
            'success': True,
            'daily': daily_metrics,
            'monthly': monthly_summary
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 400

@app.route('/api/export/excel', methods=['POST'])
def export_excel():
    """Export data to Excel file"""
    try:
        data = request.json.get('data', [])
        starting_balance = get_starting_balance()
        daily_metrics = calculate_daily_metrics(data, starting_balance)
        monthly_summary = calculate_monthly_summary(daily_metrics, starting_balance)
        
        # Create Excel file
        wb = Workbook()
        
        # Daily sheet
        ws_daily = wb.active
        ws_daily.title = "DateWise"
        
        # Headers
        headers = ['Sl', 'Date', 'Gain ($)', 'Cgain ($)', 'Loss ($)', 'Closs ($)', 
                  'Net ($)', 'Cum ($)', 'Withdrawal ($)', 'CWithdrawal ($)', 
                  'Deposit ($)', 'CDeposit ($)', 'Balance ($)']
        ws_daily.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws_daily[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Add data
        for item in daily_metrics:
            ws_daily.append([
                item.get('Sl'),
                item.get('Date'),
                item.get('Gain'),
                item.get('Cgain'),
                item.get('Loss'),
                item.get('Closs'),
                item.get('Net'),
                item.get('Cum'),
                item.get('Withdrawal'),
                item.get('CWithdrawal'),
                item.get('Deposit'),
                item.get('CDeposit'),
                item.get('Balance')
            ])
        
        # Monthly sheet
        ws_monthly = wb.create_sheet("Monthwise")
        
        headers_monthly = ['Sl', 'Month', 'Gain ($)', 'Loss ($)', 'Net ($)', 
                          'Cum ($)', 'Withdrawal ($)', 'CWithdrawal ($)', 
                          'Deposit ($)', 'CDeposit ($)', 'Balance ($)']
        ws_monthly.append(headers_monthly)
        
        for cell in ws_monthly[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        for item in monthly_summary:
            ws_monthly.append([
                item.get('Sl'),
                item.get('Month'),
                item.get('Gain'),
                item.get('Loss'),
                item.get('Net'),
                item.get('Cum'),
                item.get('Withdrawal'),
                item.get('CWithdrawal'),
                item.get('Deposit'),
                item.get('CDeposit'),
                item.get('Balance')
            ])
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'PnL_Tracker_{datetime.now().strftime("%Y%m%d")}.xlsx'
        )
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 400

@app.route('/api/import/excel', methods=['POST'])
def import_excel():
    """Import data from Excel file and save to database"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'No file provided'}), 400
        
        file = request.files['file']
        
        if file.filename.endswith('.xlsx') or file.filename.endswith('.xls'):
            df = pd.read_excel(file, sheet_name='DateWise')
        elif file.filename.endswith('.csv'):
            df = pd.read_csv(file)
        else:
            return jsonify({'success': False, 'error': 'Invalid file format'}), 400
        
        # Keep only required columns
        required_cols = ['Date', 'Gain ($)', 'Loss ($)', 'Withdrawal ($)', 'Deposit ($)']
        
        if not all(col in df.columns for col in required_cols):
            return jsonify({'success': False, 'error': 'Missing required columns'}), 400
        
        # Clear existing entries
        Entry.query.delete()
        
        # Add new entries
        for _, row in df.iterrows():
            if pd.notna(row['Date']):
                entry = Entry(
                    date=pd.to_datetime(row['Date']).date(),
                    gain=float(row['Gain ($)']) if pd.notna(row['Gain ($)']) else 0,
                    loss=float(row['Loss ($)']) if pd.notna(row['Loss ($)']) else 0,
                    withdrawal=float(row['Withdrawal ($)']) if pd.notna(row['Withdrawal ($)']) else 0,
                    deposit=float(row['Deposit ($)']) if pd.notna(row['Deposit ($)']) else 0
                )
                db.session.add(entry)
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Data imported successfully'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 400

@app.route('/api/health', methods=['GET'])
def health():
    """Health check endpoint"""
    return jsonify({'status': 'healthy', 'timestamp': datetime.now().isoformat()})

if __name__ == '__main__':
    import os
    port = int(os.environ.get('PORT', 5000))
    app.run(debug=False, host='0.0.0.0', port=port)
